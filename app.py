import streamlit as st
import pandas as pd
import io
from datetime import datetime, timedelta
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# =========================================================================
# KONFIGURASI SLA & CADENCE PER TABEL
# =========================================================================
# "cadence" menentukan sheet tujuan (Daily / Weekly / Monthly / Billing)
# "sla_label" dipakai untuk kolom SLA DATE (mis. "D+2", "M+1")
# "sla_days" dipakai untuk perhitungan TIMELINESS
#   - untuk cadence "daily": sla_days = jumlah hari (D+N)
#   - untuk cadence "monthly": sla_days = jumlah BULAN (M+N)
#
# Tabel yang tidak terdaftar di sini akan pakai DEFAULT_CONFIG di bawah.
TABLE_CONFIG = {
    # --- Tabel MONTHLY (M+1) ---
    "ih_konten_internet":            {"cadence": "monthly", "sla_label": "M+1", "sla_days": 1},
    "ih_lisrev_autocon":             {"cadence": "monthly", "sla_label": "M+1", "sla_days": 1},
    "ih_lisrev_icloud":              {"cadence": "monthly", "sla_label": "M+1", "sla_days": 1},
    "ih_lisrev_ihsmart":             {"cadence": "monthly", "sla_label": "M+1", "sla_days": 1},
    "ih_lisrev_melon":               {"cadence": "monthly", "sla_label": "M+1", "sla_days": 1},
    "ih_lisrev_movin":               {"cadence": "monthly", "sla_label": "M+1", "sla_days": 1},
    "ih_lisrev_netflix":             {"cadence": "monthly", "sla_label": "M+1", "sla_days": 1},
    "ih_lisrev_nvpr":                {"cadence": "monthly", "sla_label": "M+1", "sla_days": 1},
    "ih_lisrev_plcwifiext":          {"cadence": "monthly", "sla_label": "M+1", "sla_days": 1},
    "ih_lisrev_svod":                {"cadence": "monthly", "sla_label": "M+1", "sla_days": 1},
    "ih_tibs_lisrev_addon_usee":     {"cadence": "monthly", "sla_label": "M+1", "sla_days": 1},
    "ih_tibs_lisrev_alltv":          {"cadence": "monthly", "sla_label": "M+1", "sla_days": 1},
    "poin_redeemption":              {"cadence": "monthly", "sla_label": "M+1", "sla_days": 1},

    # --- Tabel DAILY dengan SLA khusus (bukan D+1) ---
    "ih_ga_browse_offer":            {"cadence": "daily", "sla_label": "D+4", "sla_days": 4},
    "ih_ga_order_summary":           {"cadence": "daily", "sla_label": "D+3", "sla_days": 3},
    "ih_ga_verify_otp":              {"cadence": "daily", "sla_label": "D+3", "sla_days": 3},
    "ih_tere_earning_poin":          {"cadence": "daily", "sla_label": "D+2", "sla_days": 2},
    "ih_tere_trx_0poin":             {"cadence": "daily", "sla_label": "D+2", "sla_days": 2},
    "poin_fact_detail":              {"cadence": "daily", "sla_label": "D+2", "sla_days": 2},
}

# Default untuk tabel yang tidak ada di TABLE_CONFIG -> dianggap Daily, D+1
DEFAULT_CONFIG = {"cadence": "daily", "sla_label": "D+1", "sla_days": 1}


def get_table_config(table_name: str) -> dict:
    """Ambil config (cadence, sla_label, sla_days) untuk sebuah tabel.
    Billing tetap dideteksi lewat nama, terlepas dari cadence-nya."""
    cfg = TABLE_CONFIG.get(table_name, DEFAULT_CONFIG).copy()
    if "bil" in table_name.lower() or "billing" in table_name.lower():
        cfg["sheet"] = "Billing"
    elif cfg["cadence"] == "monthly":
        cfg["sheet"] = "Monthly"
    else:
        cfg["sheet"] = "Daily"
    return cfg


# --- Fungsi bantu untuk parsing tanggal ---
def try_parse_date(value):
    """Coba ubah string ke datetime.date, kalau gagal return None."""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(value), fmt).date()
        except Exception:
            pass
    return None


# --- Fungsi utama memproses data ---
def process_data(file):
    data = file.decode("utf-8").splitlines()
    lines = [line.strip() for line in data if line.strip() and not line.startswith("||||")]

    processed_data = []
    for line in lines:
        parts = line.split("|")
        if len(parts) >= 5:
            table_name = parts[0]
            date_transaction = parts[1].replace("event_date=", "").strip()
            date_availability = parts[2]
            time_availability = parts[3]
            now_size_condition = parts[4]

            cfg = get_table_config(table_name)
            processed_data.append([
                table_name, cfg["sla_label"], date_transaction, date_availability,
                time_availability, now_size_condition
            ])

    df = pd.DataFrame(processed_data, columns=[
        "TABLE NAME", "SLA DATE", "DATE TRANSACTION", "DATE AVAILABILITY",
        "TIME AVAILABILITY", "NOW SIZE CONDITION"
    ])

    # --- Tambahkan tanggal yang hilang untuk setiap TABLE NAME ---
    # PENTING: hanya dilakukan untuk tabel DAILY. Tabel MONTHLY memang
    # cuma punya 1 baris per bulan, jadi tidak boleh di-expand jadi
    # 28-31 baris kosong (itu penyebab Monthly ketuker jadi Daily).
    all_filled = []
    for table, group in df.groupby("TABLE NAME"):
        cfg = get_table_config(table)
        group = group.copy()
        group["DATE TRANSACTION"] = pd.to_datetime(group["DATE TRANSACTION"], errors="coerce")

        if cfg["cadence"] != "daily":
            # tabel monthly/billing non-daily: jangan di-expand per hari
            group = group.sort_values("DATE TRANSACTION").reset_index(drop=True)
            all_filled.append(group)
            continue

        valid_dates = group["DATE TRANSACTION"].dropna()
        if valid_dates.empty:
            all_filled.append(group)
            continue

        year = valid_dates.dt.year.min()
        month = valid_dates.dt.month.min()

        _, last_day = calendar.monthrange(year, month)
        all_days = pd.date_range(start=f"{year}-{month:02d}-01", end=f"{year}-{month:02d}-{last_day:02d}")

        missing_days = [d for d in all_days if d not in valid_dates.values]

        if missing_days:
            new_rows = pd.DataFrame({
                "TABLE NAME": table,
                "SLA DATE": cfg["sla_label"],
                "DATE TRANSACTION": missing_days,
                "DATE AVAILABILITY": "",
                "TIME AVAILABILITY": "",
                "NOW SIZE CONDITION": ""
            })
            group = pd.concat([group, new_rows], ignore_index=True)

        group = group.sort_values("DATE TRANSACTION").reset_index(drop=True)
        all_filled.append(group)

    df = pd.concat(all_filled, ignore_index=True)

    df["DATE TRANSACTION"] = df["DATE TRANSACTION"].dt.strftime("%Y-%m-%d")

    # --- Hitung kolom COMPLETENESS ---
    df["COMPLETENESS"] = df["TIME AVAILABILITY"].apply(
        lambda x: "NOT MET" if pd.isna(x) or str(x).strip() in ["", "-"] else "MET"
    )

    # --- Hitung kolom TIMELINESS ---
    def check_timeliness(row):
        date_trans = try_parse_date(row["DATE TRANSACTION"])
        date_avail = try_parse_date(row["DATE AVAILABILITY"])
        sla_val = str(row["SLA DATE"])

        if not date_avail or str(row["DATE AVAILABILITY"]).strip() in ["", "-"]:
            return "NOT MET"
        if not date_trans:
            return "NOT MET"

        if sla_val.startswith("M+"):
            # SLA bulanan: bandingkan selisih bulan (year*12+month)
            try:
                sla_months = int(sla_val.replace("M+", ""))
            except Exception:
                sla_months = 1
            months_trans = date_trans.year * 12 + date_trans.month
            months_avail = date_avail.year * 12 + date_avail.month
            delta_months = months_avail - months_trans
            return "NOT MET" if delta_months > sla_months else "MET"

        # SLA harian (D+N), default N=1
        sla_days = 1
        if sla_val.startswith("D+"):
            try:
                sla_days = int(sla_val.replace("D+", ""))
            except Exception:
                pass

        delta = (date_avail - date_trans).days
        return "NOT MET" if delta > sla_days else "MET"

    df["TIMELINESS"] = df.apply(check_timeliness, axis=1)

    # --- Hitung kolom NOTE ---
    def check_note(row):
        if row["TIMELINESS"] == "NOT MET":
            val = str(row["NOW SIZE CONDITION"]).strip()
            if val in ["", "-"]:
                return "Source Kosong"
            else:
                return "Source Update"
        return ""

    df["NOTE"] = df.apply(check_note, axis=1)

    return df


# --- Fungsi buat workbook ---
def create_workbook(df):
    wb = Workbook()
    wb.remove(wb["Sheet"])
    sheets = {
        "Main": wb.create_sheet("Main"),
        "Daily": wb.create_sheet("Daily"),
        "Weekly": wb.create_sheet("Weekly"),
        "Monthly": wb.create_sheet("Monthly"),
        "Billing": wb.create_sheet("Billing"),
    }

    def add_table_to_sheet(ws, table_name, group):
        ws.append([f"TABLE NAME: {table_name}"])
        ws.append(list(group.columns))
        for row in group.values.tolist():
            ws.append(row)
        ws.append([])

    # Klasifikasi sheet sekarang berdasarkan TABLE_CONFIG (cadence asli
    # tabel), BUKAN berdasarkan len(group) seperti sebelumnya.
    for table_name, group in df.groupby("TABLE NAME"):
        cfg = get_table_config(table_name)
        target_sheet = sheets.get(cfg["sheet"], sheets["Main"])
        add_table_to_sheet(target_sheet, table_name, group)

    # Bersihkan teks prefix
    for sheet in sheets.values():
        for row in sheet.iter_rows():
            if row[0].value and "TABLE NAME:" in str(row[0].value):
                row[0].value = row[0].value.replace("TABLE NAME: ", "")

    return wb


# --- Format warna Excel ---
def format_excel_with_feeds(wb):
    first_header_fill = PatternFill(start_color="3C7D22", end_color="3C7D22", fill_type="solid")
    second_header_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        max_row = sheet.max_row
        max_col = 9

        for col_num, col_cells in enumerate(sheet.columns, start=1):
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
            sheet.column_dimensions[get_column_letter(col_num)].width = max_length + 2

        row = 1
        while row <= max_row:
            if sheet.cell(row=row, column=1).value:
                sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
                cell = sheet.cell(row=row, column=1)
                cell.fill = first_header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                row += 1
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell.fill = second_header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                row += 1
                while row <= max_row and sheet.cell(row=row, column=1).value:
                    for col in range(1, max_col + 1):
                        cell = sheet.cell(row=row, column=col)
                        cell.border = thin_border
                    row += 1
            else:
                row += 1
    return wb


# --- Save ke BytesIO ---
def save_workbook_to_bytes(wb):
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# --- Streamlit UI ---
st.title("📊 Data Processing and Excel Export with Logic Evaluation")
st.write("Upload a `data.txt` file to process and see MET/NOT MET logic applied automatically.")

uploaded_file = st.file_uploader("Choose a file", type="txt")
if uploaded_file:
    df = process_data(uploaded_file.read())
    st.write("### Processed Data")
    st.dataframe(df)

    if st.button("Generate Excel File"):
        workbook = create_workbook(df)
        formatted_workbook = format_excel_with_feeds(workbook)
        excel_file = save_workbook_to_bytes(formatted_workbook)
        st.download_button(
            label="📥 Download Evaluated Excel File",
            data=excel_file,
            file_name="output_evaluated.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
